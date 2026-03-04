import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import zipfile
import io
import os
from typing import Dict, List, Tuple, Optional
import re
import hmac

# Set page configuration - removed page_layout parameter
st.set_page_config(
    page_title="Course Performance Analyzer",
    initial_sidebar_state="expanded"
)

class CourseAnalyzer:
    """Main class for course analysis operations"""
    
    def __init__(self):
        self.department_map = {
            'CS': 'Computer Science',
            'IS': 'Information Systems',
            'IT': 'Information Technology'
        }

    def validate_file_structure(self, df: pd.DataFrame) -> Tuple[bool, List[str]]:
        """
        Validate the file structure according to requirements
        Returns: (is_valid, list_of_errors)
        """
        errors = []
        
        try:
            # Check if file has enough rows and columns
            if len(df) < 7 or len(df.columns) < 4:
                errors.append("File does not follow the required template format")
                return False, errors
            
            # === CHECK 1: Header cells in column A (rows 1-4) must be EXACT ===
            expected_headers = {
                0: 'Course Title',   # Row 1
                1: 'Cours Code',    # Row 2
                2: 'Section',        # Row 3
                3: 'Instructor Name' # Row 4
            }
            
            for row_idx, expected in expected_headers.items():
                cell_value = df.iloc[row_idx, 0] if len(df.columns) > 0 else ""
                cell_text = clean_text(cell_value).replace(':', '').strip()
                
                if cell_text != expected:
                    errors.append("File does not follow the required template format")
                    return False, errors
            
            # === CHECK 2: C1-C4 must have data ===
            for row_idx in range(4):
                cell_value = df.iloc[row_idx, 2] if len(df.columns) > 2 else ""
                cell_text = clean_text(cell_value)
                
                if not cell_text or cell_text.strip() == "":
                    errors.append("File does not follow the required template format")
                    return False, errors
            
            # === CHECK 3: Header row at row 6 must be EXACT ===
            header_row_idx = 5  # Row 6
            
            # Get header values
            col_a = clean_text(df.iloc[header_row_idx, 0]) if len(df.columns) > 0 else ""
            col_b = clean_text(df.iloc[header_row_idx, 1]) if len(df.columns) > 1 else ""
            col_c = clean_text(df.iloc[header_row_idx, 2]) if len(df.columns) > 2 else ""
            col_d = clean_text(df.iloc[header_row_idx, 3]) if len(df.columns) > 3 else ""
            
            # Must match exactly
            if col_a != "No" or col_b != "ID" or col_c != "Name" or col_d != "Midterm":
                errors.append("File does not follow the required template format")
                return False, errors
            
            # === CHECK 4: Must have at least one student record with data in all columns ===
            has_valid_record = False
            for check_row in range(header_row_idx + 1, len(df)):
                # Check if row has data in all required columns
                if (len(df.columns) >= 4 and
                    pd.notna(df.iloc[check_row, 0]) and str(df.iloc[check_row, 0]).strip() != "" and
                    pd.notna(df.iloc[check_row, 1]) and str(df.iloc[check_row, 1]).strip() != "" and
                    pd.notna(df.iloc[check_row, 2]) and str(df.iloc[check_row, 2]).strip() != "" and
                    pd.notna(df.iloc[check_row, 3])):
                    
                    # Try to convert midterm to number
                    try:
                        float(df.iloc[check_row, 3])
                        has_valid_record = True
                        break
                    except:
                        continue
            
            if not has_valid_record:
                errors.append("File does not follow the required template format")
                return False, errors
            
        except Exception as e:
            errors.append("File does not follow the required template format")
            return False, errors
        
        return True, []

    def extract_course_info(self, df: pd.DataFrame) -> Dict:
        """Extract course information from header cells following exact template format"""
        try:
            # Template expects course info in specific positions:
            # Row 1 (index 0), Column C: Course Title
            # Row 2 (index 1), Column C: Course Code  
            # Row 3 (index 2), Column C: Section
            # Row 4 (index 3), Column C: Instructor Name
            
            course_title = ""
            course_code = ""
            section = ""
            instructor = ""
            
            # Only check rows 0-3 (rows 1-4 in Excel)
            # Only check column C (index 2)
            if len(df) > 0 and len(df.columns) > 2:
                # Row 1: Course Title
                if 0 < len(df):
                    course_title = clean_text(df.iloc[0, 2]) if pd.notna(df.iloc[0, 2]) else ""
                
                # Row 2: Course Code
                if 1 < len(df):
                    course_code = clean_text(df.iloc[1, 2]) if pd.notna(df.iloc[1, 2]) else ""
                
                # Row 3: Section
                if 2 < len(df):
                    section = clean_text(df.iloc[2, 2]) if pd.notna(df.iloc[2, 2]) else ""
                
                # Row 4: Instructor Name
                if 3 < len(df):
                    instructor = clean_text(df.iloc[3, 2]) if pd.notna(df.iloc[3, 2]) else ""
            
            return {
                'title': course_title,
                'code': course_code,
                'section': section,
                'instructor': instructor
            }
        except Exception as e:
            return {'title': '', 'code': '', 'section': '', 'instructor': ''}

    def extract_midterm_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Extract midterm data starting from row 7"""
        try:
            # Start from row 7 (index 6) to end
            data = df.iloc[6:].copy()
            
            # Reset index and rename columns
            if len(data.columns) >= 4:
                data.columns = ['No', 'ID', 'Name', 'Midterm'] + list(data.columns[4:])
            else:
                # Handle case with fewer columns
                cols = ['No', 'ID', 'Name', 'Midterm']
                data.columns = cols[:len(data.columns)]
            
            # Clean up data
            if 'No' in data.columns:
                data = data[['No', 'ID', 'Name', 'Midterm']].copy() if all(col in data.columns for col in ['No', 'ID', 'Name', 'Midterm']) else data
            else:
                # If columns don't match, return empty DataFrame
                return pd.DataFrame()
            
            # Remove rows where No is NaN or empty (these are truly empty rows)
            data = data[data['No'].notna() & (data['No'] != '')]
            
            # Convert No to numeric to identify valid student rows
            data['No_Numeric'] = pd.to_numeric(data['No'], errors='coerce')
            
            # A student exists ONLY if they have a valid ID and Name
            # Check if ID and Name are filled (not empty)
            data['ID_Valid'] = data['ID'].notna() & (data['ID'].astype(str).str.strip() != '')
            data['Name_Valid'] = data['Name'].notna() & (data['Name'].astype(str).str.strip() != '')
            
            # Student exists only if they have valid ID AND Name
            data['Student_Exists'] = data['ID_Valid'] & data['Name_Valid']
            
            # Convert Midterm to numeric for existing students
            data['Midterm_Value'] = pd.to_numeric(data['Midterm'], errors='coerce')
            
            # For existing students, check if they were absent (missing midterm)
            # For non-existing students, they shouldn't be counted at all
            data['Is_Absent'] = False
            data.loc[data['Student_Exists'], 'Is_Absent'] = (
                data.loc[data['Student_Exists'], 'Midterm'].isna() | 
                (data.loc[data['Student_Exists'], 'Midterm'].astype(str).str.strip() == '') | 
                data.loc[data['Student_Exists'], 'Midterm_Value'].isna()
            )
            
            # Get valid scores only for existing students
            data['Score'] = np.nan
            data.loc[data['Student_Exists'] & ~data['Is_Absent'], 'Score'] = data.loc[data['Student_Exists'] & ~data['Is_Absent'], 'Midterm_Value']
            
            # Filter to only include existing students
            existing_students = data[data['Student_Exists']].copy()
            
            # Calculate statistics
            total_students = len(existing_students)
            absent_count = existing_students['Is_Absent'].sum()
            valid_scores = existing_students[~existing_students['Is_Absent']]['Score'].dropna().tolist()
            
            # Store for debugging if needed
            existing_students['total_students'] = total_students
            existing_students['absent_count'] = absent_count
            
            return existing_students
            
        except Exception as e:
            st.error(f"Error extracting midterm data: {str(e)}")
            return pd.DataFrame()

    def get_department(self, course_code: str) -> str:
        """Determine department based on course code containing department codes"""
        course_code_upper = str(course_code).upper()  # Convert to string and uppercase for case-insensitive matching
        
        for code, dept in self.department_map.items():
            if code in course_code_upper:  # Check if code is contained in course_code, not just at start
                return dept
        return 'Other'
    
    

    def aggregate_courses(self, courses_data: Dict[str, Dict]) -> Dict[str, Dict]:
        """
        Aggregate multiple sections of the same course
        Key is tuple of (course_title, course_code)
        """
        aggregated = {}
        
        for filename, data in courses_data.items():
            if not data['valid']:
                continue
                
            course_info = data['course_info']
            midterm_data = data['midterm_data']
            
            if midterm_data.empty:
                continue
                
            # Create course key
            course_key = (course_info['title'], course_info['code'])
            
            if course_key not in aggregated:
                aggregated[course_key] = {
                    'title': course_info['title'],
                    'code': course_info['code'],
                    'sections': [],
                    'all_scores': [],
                    'all_students': [],
                    'absent_count': 0,
                    'total_students': 0,
                    'department': self.get_department(course_info['code'])
                }
            
            # Add this section's data
            valid_scores = midterm_data[~midterm_data['Is_Absent']]['Score'].dropna().tolist()
            absent_count = midterm_data['Is_Absent'].sum()
            total_students = len(midterm_data)  # Now this only counts existing students
            
            aggregated[course_key]['sections'].append(course_info['section'])
            aggregated[course_key]['all_scores'].extend(valid_scores)
            aggregated[course_key]['absent_count'] += absent_count
            aggregated[course_key]['total_students'] += total_students
            aggregated[course_key]['all_students'].extend(midterm_data.to_dict('records'))
        
        return aggregated
    
    def analyze_failure_rates(self, aggregated_courses: Dict, min_students: int = 5, 
                            failure_threshold: float = 50, failure_score: float = 12) -> pd.DataFrame:
        """Analyze courses with high failure rates (> threshold% scored <= failure_score)"""
        results = []
        
        for (title, code), data in aggregated_courses.items():
            if data['total_students'] < min_students:
                continue
                
            scores = data['all_scores']
            if len(scores) == 0:
                continue
                
            students_failed = sum(1 for s in scores if s <= failure_score)
            percentage = (students_failed / len(scores)) * 100
            
            if percentage > failure_threshold:
                results.append({
                    'Course Title': data['title'],
                    'Course Code': data['code'],
                    'Department': data['department'],
                    'Total Students': data['total_students'],
                    'Absent Students': data['absent_count'],
                    f'Students with Scores ≤ {failure_score}': students_failed,
                    'Total Valid Scores': len(scores),
                    'Failure Percentage': round(percentage, 2)
                })
        
        return pd.DataFrame(results)
    
    def analyze_success_rates(self, aggregated_courses: Dict, min_students: int = 5,
                            success_threshold: float = 50, success_score: float = 17) -> pd.DataFrame:
        """Analyze courses with high success rates (> threshold% scored >= success_score)"""
        results = []
        
        for (title, code), data in aggregated_courses.items():
            if data['total_students'] < min_students:
                continue
                
            scores = data['all_scores']
            if len(scores) == 0:
                continue
                
            students_success = sum(1 for s in scores if s >= success_score)
            percentage = (students_success / len(scores)) * 100
            
            if percentage > success_threshold:
                results.append({
                    'Course Title': data['title'],
                    'Course Code': data['code'],
                    'Department': data['department'],
                    'Total Students': data['total_students'],
                    'Absent Students': data['absent_count'],
                    f'Students with Scores ≥ {success_score}': students_success,
                    'Total Valid Scores': len(scores),
                    'Success Percentage': round(percentage, 2)
                })
        
        return pd.DataFrame(results)

def process_uploaded_file(file, analyzer: CourseAnalyzer):
    """Process a single uploaded file with robust handling for various Excel formats"""
    try:
        # Get filename and handle Unicode properly
        filename = getattr(file, 'name', 'unknown_file')
        if isinstance(filename, bytes):
            try:
                filename = filename.decode('utf-8', errors='ignore')
            except:
                filename = str(filename)
        
        # Reset file pointer to beginning
        if hasattr(file, 'seek'):
            file.seek(0)
        
        df = None
        errors = []
        
        # Save the file content temporarily
        file_content = file.read()
        
        # Method 1: Try with openpyxl directly (best for .xlsx)
        try:
            from openpyxl import load_workbook
            
            file_bytes = io.BytesIO(file_content)
            file_bytes.seek(0)
            
            # Try to load workbook with different options
            try:
                wb = load_workbook(file_bytes, read_only=True, data_only=True)
            except:
                # If read_only fails, try without it
                file_bytes.seek(0)
                wb = load_workbook(file_bytes, data_only=True)
            
            sheet = wb.active
            
            # Convert to DataFrame
            data = []
            for row in sheet.iter_rows(values_only=True):
                # Convert any cell values to appropriate types
                row_data = []
                for cell in row:
                    if cell is None:
                        row_data.append('')
                    elif isinstance(cell, (int, float)):
                        row_data.append(cell)
                    else:
                        # Convert to string and handle encoding
                        try:
                            if isinstance(cell, bytes):
                                cell = cell.decode('utf-8', errors='ignore')
                            else:
                                cell = str(cell)
                            # Clean up the string
                            cell = cell.strip()
                            row_data.append(cell)
                        except:
                            row_data.append('')
                data.append(row_data)
            
            if data:
                df = pd.DataFrame(data)
                # print(f"Successfully read with openpyxl direct: {filename}")
            
            wb.close()
        except Exception as e:
            errors.append(f"openpyxl direct failed: {str(e)}")
        
        # Method 2: Try with pandas + openpyxl
        if df is None:
            try:
                file_bytes = io.BytesIO(file_content)
                file_bytes.seek(0)
                df = pd.read_excel(file_bytes, header=None, engine='openpyxl')
                # print(f"Successfully read with pandas+openpyxl: {filename}")
            except Exception as e:
                errors.append(f"pandas+openpyxl failed: {str(e)}")
        
        # Method 3: Try with pandas + xlrd (for older .xls files)
        if df is None:
            try:
                file_bytes = io.BytesIO(file_content)
                file_bytes.seek(0)
                df = pd.read_excel(file_bytes, header=None, engine='xlrd')
                # print(f"Successfully read with pandas+xlrd: {filename}")
            except Exception as e:
                errors.append(f"pandas+xlrd failed: {str(e)}")
        
        # Method 4: Try with pandas auto-detect
        if df is None:
            try:
                file_bytes = io.BytesIO(file_content)
                file_bytes.seek(0)
                df = pd.read_excel(file_bytes, header=None)
                # print(f"Successfully read with pandas auto: {filename}")
            except Exception as e:
                errors.append(f"pandas auto failed: {str(e)}")
        
        if df is None:
            error_msg = "\n".join(errors)
            return {
                'filename': filename,
                'valid': False,
                'errors': errors + ["Could not read Excel file with any method"],
                'course_info': None,
                'midterm_data': None
            }
        
        # Clean up the DataFrame - handle any byte strings and Arabic text
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].apply(lambda x: clean_text(x))
        
        # # === DEBUG: Print first 10 rows to see what's in the file ===
        # st.write(f"=== DEBUG: First 10 rows of {filename} ===")
        # for i in range(min(10, len(df))):
        #     row_values = []
        #     for j in range(min(4, len(df.columns))):
        #         val = df.iloc[i, j] if j < len(df.columns) else ""
        #         row_values.append(str(val)[:20])
        #     st.write(f"Row {i+1}: {row_values}")
        
        # Validate structure
        is_valid, validation_errors = analyzer.validate_file_structure(df)
        
        # st.write(f"Validation result: {is_valid}")
        # if validation_errors:
        #     st.write(f"Validation errors: {validation_errors}")
        
        if not is_valid:
            return {
                'filename': filename,
                'valid': False,
                'errors': validation_errors,
                'course_info': None,
                'midterm_data': None,
                'raw_df': df
            }
        
        # Extract course info and midterm data
        course_info = analyzer.extract_course_info(df)
        midterm_data = analyzer.extract_midterm_data(df)
        
        # st.write(f"Course info extracted: {course_info}")
        # st.write(f"Midterm data shape: {midterm_data.shape if not midterm_data.empty else 'Empty'}")
        
        return {
            'filename': filename,
            'valid': True,
            'errors': [],
            'course_info': course_info,
            'midterm_data': midterm_data,
            'raw_df': df
        }
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return {
            'filename': getattr(file, 'name', 'unknown_file'),
            'valid': False,
            'errors': [f"Error processing file: {str(e)}"],
            'course_info': None,
            'midterm_data': None
        }

def clean_text(x):
    """Clean and decode text properly"""
    if pd.isna(x):
        return ""
    if isinstance(x, (int, float)):
        return str(x)
    if isinstance(x, bytes):
        try:
            # Try UTF-8 first
            return x.decode('utf-8', errors='ignore').strip()
        except:
            try:
                # Try Arabic-specific encodings
                return x.decode('cp1256', errors='ignore').strip()
            except:
                # Last resort
                return x.decode('latin-1', errors='ignore').strip()
    # Convert to string and clean
    text = str(x).strip()
    # Remove any weird characters
    text = ''.join(char for char in text if ord(char) >= 32 or char == '\n')
    return text

def create_failure_chart(df: pd.DataFrame, failure_score: float, title: str = "High Failure Rate Courses"):
    """Create interactive bar chart for failure analysis"""
    if df.empty:
        return None
    
    fig = make_subplots(
        rows=2, cols=1,
        subplot_titles=(f"Students with Scores ≤ {failure_score}", "Failure Percentage"),
        vertical_spacing=0.15
    )
    
    # Get the dynamic column name
    failed_col = f'Students with Scores ≤ {failure_score}'
    
    # Add bar chart for student counts
    fig.add_trace(
        go.Bar(
            x=df['Course Code'],
            y=df[failed_col],
            name='Failed Students',
            marker_color='crimson',
            text=df[failed_col],
            textposition='outside',
            hovertemplate='<b>%{x}</b><br>Failed Students: %{y}<br>Total: %{customdata[0]}<br>Percentage: %{customdata[1]}%<extra></extra>',
            customdata=df[['Total Valid Scores', 'Failure Percentage']].values
        ),
        row=1, col=1
    )
    
    # Add bar chart for percentages
    fig.add_trace(
        go.Bar(
            x=df['Course Code'],
            y=df['Failure Percentage'],
            name='Percentage',
            marker_color='darkorange',
            text=df['Failure Percentage'].apply(lambda x: f'{x:.1f}%'),
            textposition='outside',
            hovertemplate='<b>%{x}</b><br>Failure Rate: %{y:.1f}%<extra></extra>'
        ),
        row=2, col=1
    )
    
    fig.update_layout(
        title=title,
        showlegend=False,
        height=600,
        hovermode='x unified'
    )
    
    fig.update_xaxes(title_text="Course Code", row=2, col=1)
    fig.update_yaxes(title_text="Number of Students", row=1, col=1)
    fig.update_yaxes(title_text="Percentage (%)", row=2, col=1)
    
    return fig

def create_success_chart(df: pd.DataFrame, success_score: float, title: str = "High Success Rate Courses"):
    """Create interactive bar chart for success analysis"""
    if df.empty:
        return None
    
    fig = make_subplots(
        rows=2, cols=1,
        subplot_titles=(f"Students with Scores ≥ {success_score}", "Success Percentage"),
        vertical_spacing=0.15
    )
    
    # Get the dynamic column name
    success_col = f'Students with Scores ≥ {success_score}'
    
    # Add bar chart for student counts
    fig.add_trace(
        go.Bar(
            x=df['Course Code'],
            y=df[success_col],
            name='Successful Students',
            marker_color='forestgreen',
            text=df[success_col],
            textposition='outside',
            hovertemplate='<b>%{x}</b><br>Successful Students: %{y}<br>Total: %{customdata[0]}<br>Percentage: %{customdata[1]}%<extra></extra>',
            customdata=df[['Total Valid Scores', 'Success Percentage']].values
        ),
        row=1, col=1
    )
    
    # Add bar chart for percentages
    fig.add_trace(
        go.Bar(
            x=df['Course Code'],
            y=df['Success Percentage'],
            name='Percentage',
            marker_color='goldenrod',
            text=df['Success Percentage'].apply(lambda x: f'{x:.1f}%'),
            textposition='outside',
            hovertemplate='<b>%{x}</b><br>Success Rate: %{y:.1f}%<extra></extra>'
        ),
        row=2, col=1
    )
    
    fig.update_layout(
        title=title,
        showlegend=False,
        height=600,
        hovermode='x unified'
    )
    
    fig.update_xaxes(title_text="Course Code", row=2, col=1)
    fig.update_yaxes(title_text="Number of Students", row=1, col=1)
    fig.update_yaxes(title_text="Percentage (%)", row=2, col=1)
    
    return fig

def check_auth() -> None:
    """Shared username/password auth using Streamlit secrets."""
    if st.session_state.get("auth_ok", False):
        return

    st.title("🔐 Login")

    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Log in")

    if submitted:
        user_ok = hmac.compare_digest(username, str(st.secrets.get("APP_USERNAME", "")))
        pass_ok = hmac.compare_digest(password, str(st.secrets.get("APP_PASSWORD", "")))

        if user_ok and pass_ok:
            st.session_state["auth_ok"] = True
            st.rerun()
        else:
            st.error("Invalid username or password.")
            st.stop()

    # Block the rest of the app until authenticated
    st.stop()

def main():
    check_auth()
    # Set page title and configure for RTL if needed
    st.set_page_config(
        page_title="Course Performance Analyzer",
        initial_sidebar_state="expanded"
    )
    
    # Add CSS for better Arabic text display
    st.markdown("""
    <style>
    /* Support for Arabic text */
    .arabic-text {
        font-family: 'Arial', 'Helvetica', sans-serif;
        direction: rtl;
        text-align: right;
    }
    </style>
    """, unsafe_allow_html=True)

    st.title("📊 Course Performance Analyzer")
    st.markdown("Upload Excel files to analyze course performance based on midterm marks")
    
    # Initialize analyzer
    analyzer = CourseAnalyzer()
    
    # Sidebar for configuration
    with st.sidebar:

        st.header("📁 File Upload")
        
        # File upload options
        upload_option = st.radio(
            "Choose upload method:",
            ["Upload individual files", "Upload ZIP file"]
        )

        uploaded_files = []
        if upload_option == "Upload individual files":
            files = st.file_uploader(
                "Select Excel files",
                type=['xlsx', 'xls'],
                accept_multiple_files=True
            )
            if files:
                uploaded_files.extend(files)
        else:
            zip_file = st.file_uploader(
                "Select ZIP file",
                type=['zip']
            )
            if zip_file:
                with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                    for zip_info in zip_ref.infolist():
                        file_name = zip_info.filename
                        
                        # Get the raw bytes of the filename if possible
                        if hasattr(zip_info, 'raw_filename'):
                            raw_bytes = zip_info.raw_filename
                        else:
                            # If raw_filename not available, try to encode the string back to bytes
                            try:
                                # Try to encode with cp437 (ZIP default encoding)
                                raw_bytes = file_name.encode('cp437')
                            except:
                                raw_bytes = file_name.encode('utf-8', errors='ignore')
                        
                        # Try to decode with the correct encoding
                        try:
                            # Try UTF-8 first
                            decoded_name = raw_bytes.decode('utf-8')
                        except:
                            try:
                                # Try Arabic Windows encoding
                                decoded_name = raw_bytes.decode('cp1256')
                            except:
                                try:
                                    # Try Latin-1 as fallback
                                    decoded_name = raw_bytes.decode('latin-1')
                                except:
                                    # If all else fails, use the original
                                    decoded_name = file_name
                        
                        # Get just the filename without path
                        base_name = os.path.basename(decoded_name)

                        # Only process valid Excel files
                        if (base_name.lower().endswith(('.xlsx', '.xls')) and 
                            '~$' not in base_name and  # Skip any file with ~$ in the name
                            not base_name.startswith(('._')) and
                            not base_name.startswith('.') and
                            not base_name == 'Thumbs.db' and
                            not base_name == '.DS_Store'):
                            
                            with zip_ref.open(file_name) as file:
                                file_content = file.read()
                                file_obj = io.BytesIO(file_content)
                                file_obj.name = base_name
                                file_obj.seek(0)
                                uploaded_files.append(file_obj)

        st.markdown("---")
        st.header("⚙️ Configuration")
        
        # Minimum student count filter
        min_students = st.number_input(
            "Minimum Students per Course",
            min_value=1,
            max_value=50,
            value=5,
            help="Exclude courses with fewer than this many total students"
        )

        st.markdown("---")
        st.header("📊 Analysis Thresholds")
        
        failure_score = st.number_input(
            "Failure Score (≤)",
            min_value=0,
            max_value=20,
            value=12,
            help="Students with scores less than or equal to this value are considered failing"
        )
        
        success_score = st.number_input(
            "Success Score (≥)",
            min_value=0,
            max_value=20,
            value=17,
            help="Students with scores greater than or equal to this value are considered successful"
        )
        
        # Validate that failure score is less than success score
        if failure_score >= success_score:
            st.error("❌ Failure score must be less than success score!")
            st.stop()

        # Failure rate threshold
        failure_threshold = st.slider(
            "High Failure Rate Threshold (%)",
            min_value=0,
            max_value=100,
            value=50,
            help="Courses with failure percentage above this value will be shown"
        )
        
        # Success rate threshold
        success_threshold = st.slider(
            "High Success Rate Threshold (%)",
            min_value=0,
            max_value=100,
            value=50,
            help="Courses with success percentage above this value will be shown"
        )

    # Main content area
    if uploaded_files:
        st.header("📋 Validation Results")
        
        # Process all files
        all_results = []
        validation_passed = []
        validation_failed = []
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for i, file in enumerate(uploaded_files):
            status_text.text(f"Processing file {i+1} of {len(uploaded_files)}: {getattr(file, 'name', 'unknown')}")
            result = process_uploaded_file(file, analyzer)
            all_results.append(result)
            
            if result['valid']:
                validation_passed.append(result)
            else:
                validation_failed.append(result)
            
            progress_bar.progress((i + 1) / len(uploaded_files))
        
        status_text.text("Processing complete!")
        
        # Display validation summary
        col1, col2 = st.columns(2)
        with col1:
            st.success(f"✅ Passed: {len(validation_passed)} files")
        with col2:
            st.error(f"❌ Failed: {len(validation_failed)} files")
        
        # Show failed files
        if validation_failed:
            with st.expander("View Failed Files Details"):
                for result in validation_failed:
                    st.error(f"**{result['filename']}**")
                    for error in result['errors']:
                        st.write(f"- {error}")
                    st.markdown("---")

        # Aggregate and analyze if there are valid files
        if validation_passed:
            st.header("📊 Course Analysis")
            
            # Aggregate courses
            courses_data = {r['filename']: r for r in validation_passed}
            aggregated = analyzer.aggregate_courses(courses_data)
            
            if not aggregated:
                st.warning("No valid course data found in the uploaded files.")
                return
            
            # Get unique departments for dropdown
            departments = set()
            for data in aggregated.values():
                departments.add(data['department'])
            departments_list = sorted(list(departments))
            departments_list.insert(0, "All")  # Add "All" option at the beginning
            
            # Department dropdown menu
            selected_dept = st.selectbox(
                "🔍 Filter by Department:",
                departments_list,
                help="Select a department to filter the analysis results"
            )
            
            # Filter courses based on selected department
            if selected_dept == "All":
                filtered_aggregated = aggregated
            else:
                filtered_aggregated = {
                    k: v for k, v in aggregated.items() 
                    if v['department'] == selected_dept
                }
            
            # Create tabs for different analyses
            tab1, tab2, tab3 = st.tabs([
                f"🔴 High Failure Rate Courses (> {failure_threshold}% ≤ {failure_score})",
                f"🟢 High Success Rate Courses (> {success_threshold}% ≥ {success_score})",
                "📊 Course Details"
            ])
            
            with tab1:
                st.subheader(f"Courses with >{failure_threshold}% Students Scoring ≤{failure_score}")
                
                # Perform failure analysis on filtered data with user thresholds
                failure_df = analyzer.analyze_failure_rates(
                    filtered_aggregated, 
                    min_students, 
                    failure_threshold, 
                    failure_score
                )
                
                if not failure_df.empty:
                    # Display filter info
                    if selected_dept != "All":
                        st.info(f"Showing results for **{selected_dept}** department")

                    # Add serial number column (starting from 1)
                    failure_df_with_index = failure_df.reset_index(drop=True)
                    failure_df_with_index.index = failure_df_with_index.index + 1  # Start from 1
                    failure_df_with_index.index.name = "#"
                    
                    # Display table with serial numbers
                    display_cols = ['Course Title', 'Course Code', 'Department', 'Total Students', 
                                  'Absent Students', f'Students with Scores ≤ {failure_score}', 
                                  'Failure Percentage']
                    st.dataframe(
                        failure_df_with_index[display_cols],
                        width='stretch'
                    )                    
                    
                    # Create visualization
                    fig = create_failure_chart(failure_df, failure_score)
                    if fig:
                        st.plotly_chart(fig, width='stretch')
                    
                    # Download option
                    csv = failure_df.to_csv(index=False)
                    st.download_button(
                        label="📥 Download Failure Analysis CSV",
                        data=csv,
                        file_name=f"failure_analysis_{selected_dept.replace(' ', '_')}.csv",
                        mime="text/csv"
                    )
                else:
                    if selected_dept != "All":
                        st.info(f"No high failure rate courses found in the **{selected_dept}** department")
                    else:
                        st.info("No courses found with high failure rates meeting the criteria")
            
            with tab2:
                st.subheader(f"Courses with >{success_threshold}% Students Scoring ≥{success_score}")
                
                # Perform success analysis on filtered data with user thresholds
                success_df = analyzer.analyze_success_rates(
                    filtered_aggregated, 
                    min_students, 
                    success_threshold, 
                    success_score
                )
                
                if not success_df.empty:
                    # Display filter info
                    if selected_dept != "All":
                        st.info(f"Showing results for **{selected_dept}** department")
                    
                     # Add serial number column (starting from 1)
                    success_df_with_index = success_df.reset_index(drop=True)
                    success_df_with_index.index = success_df_with_index.index + 1  # Start from 1
                    success_df_with_index.index.name = "#"                   

                    # Display table
                    display_cols = ['Course Title', 'Course Code', 'Department', 'Total Students',
                                  'Absent Students', f'Students with Scores ≥ {success_score}',
                                  'Success Percentage']

                    st.dataframe(
                        success_df_with_index[display_cols],
                        width='stretch'
                    )

                    # Create visualization
                    fig = create_success_chart(success_df, success_score)
                    if fig:
                        st.plotly_chart(fig, width='stretch')
                    
                    # Download option
                    csv = success_df.to_csv(index=False)
                    st.download_button(
                        label="📥 Download Success Analysis CSV",
                        data=csv,
                        file_name=f"success_analysis_{selected_dept.replace(' ', '_')}.csv",
                        mime="text/csv"
                    )
                else:
                    if selected_dept != "All":
                        st.info(f"No high success rate courses found in the **{selected_dept}** department")
                    else:
                        st.info("No courses found with high success rates meeting the criteria")

            with tab3:
                st.subheader("Course Details & Statistics")
                
                # Create a list of available courses for dropdown
                course_options = []
                course_mapping = {}
                
                for (title, code), data in aggregated.items():
                    display_name = f"{code} - {title} (Section: {', '.join(data['sections']) if data['sections'] else 'N/A'})"
                    course_options.append(display_name)
                    course_mapping[display_name] = {
                        'title': title,
                        'code': code,
                        'data': data
                    }
                
                if course_options:
                    selected_course = st.selectbox(
                        "🔍 Select a course to view details:",
                        course_options,
                        help="Choose a course to see detailed statistics and marks distribution"
                    )
                    
                    if selected_course:
                        course_info = course_mapping[selected_course]
                        course_data = course_info['data']
                        
                        # 1. Course Information - Displayed as text
                        st.markdown("### 📋 Course Information")
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown(f"**Course Code:** {course_info['code']}")
                            st.markdown(f"**Course Title:** {course_info['title']}")
                        with col2:
                            st.markdown(f"**Department:** {course_data['department']}")
                            st.markdown(f"**Sections:** {', '.join(course_data['sections']) if course_data['sections'] else 'N/A'}")
                        
                        st.markdown("---")
                        
                        # 2. Statistics
                        st.markdown("### 📊 Statistics")
                        
                        # Calculate statistics
                        all_scores = course_data['all_scores']
                        total_students = course_data['total_students']
                        absent_count = course_data['absent_count']
                        present_students = total_students - absent_count
                        
                        if all_scores:
                            avg_score = sum(all_scores) / len(all_scores)
                            max_score = max(all_scores)
                            min_score = min(all_scores)
                            median_score = sorted(all_scores)[len(all_scores)//2]
                            
                            # Display statistics in a clean layout
                            col1, col2, col3, col4 = st.columns(4)
                            with col1:
                                st.metric("Total Students", total_students)
                            with col2:
                                st.metric("Present", present_students)
                            with col3:
                                st.metric("Absent", absent_count)
                            with col4:
                                st.metric("Valid Scores", len(all_scores))
                            
                            col1, col2, col3, col4 = st.columns(4)
                            with col1:
                                st.metric("Average Score", f"{avg_score:.2f}")
                            with col2:
                                st.metric("Highest Score", f"{max_score:.2f}")
                            with col3:
                                st.metric("Lowest Score", f"{min_score:.2f}")
                            with col4:
                                st.metric("Median Score", f"{median_score:.2f}")
                            
                            # 4. Marks Distribution
                            st.markdown("### 📊 Marks Distribution")
                            
                            # Create histogram
                            fig_hist = go.Figure()
                            fig_hist.add_trace(go.Histogram(
                                x=all_scores,
                                nbinsx=20,
                                marker_color='royalblue',
                                opacity=0.7,
                                name='Marks'
                            ))
                            
                            # Add vertical lines for thresholds
                            fig_hist.add_vline(x=failure_score, line_dash="dash", line_color="red",
                                              annotation_text=f"Failure ≤{failure_score}", 
                                              annotation_position="top right")
                            fig_hist.add_vline(x=success_score, line_dash="dash", line_color="green",
                                              annotation_text=f"Success ≥{success_score}", 
                                              annotation_position="top left")
                            
                            fig_hist.update_layout(
                                title=f"Marks Distribution - {course_info['code']}",
                                xaxis_title="Marks",
                                yaxis_title="Number of Students",
                                height=400,
                                showlegend=False
                            )
                            st.plotly_chart(fig_hist, width='stretch')
                            
                        else:
                            st.warning("No valid score data available for this course")
                else:
                    st.info("No courses available for analysis")          


            # Overall summary
            st.header("📈 Summary Statistics")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("Files Processed", len(validation_passed))
            with col2:
                st.metric("Courses in Filter", len(filtered_aggregated))
            with col3:
                valid_courses = len([d for d in filtered_aggregated.values() if d['total_students'] >= min_students])
                st.metric("Courses Analyzed", valid_courses)
            
            # Add current thresholds info
            st.caption(f"📊 Current thresholds: Failure >{failure_threshold}% (≤{failure_score}) | Success >{success_threshold}% (≥{success_score})")
            # Optional: Add a small summary of the filter
            if selected_dept != "All":
                st.caption(f"📊 Currently viewing: **{selected_dept}** department only")
            else:
                st.caption("📊 Currently viewing: **All departments**")
    
    else:
        # Welcome message when no files are uploaded
        st.info("👈 Please upload Excel files or a ZIP file containing Excel files to begin analysis")
        
        # Example format
        with st.expander("📝 Expected File Format"):
            st.markdown("""
            **Header Section (Rows 1-4):**
            - A1: 'Course Title'
            - A2: 'Course Code'
            - A3: 'Section'
            - A4: 'Instructor Name'
            - C1-C4: Must be filled with data
            
            **Data Header Row (Row 6):**
            - A6: 'No'
            - B6: 'ID'
            - C6: 'Name'
            - D6: 'Midterm'
            
            **Data Section (Row 7 onwards):**
            - Student records with midterm marks in column D
            """)

if __name__ == "__main__":
    main()    